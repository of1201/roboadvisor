#!/usr/bin/env python
# coding: utf-8

# # Robo advisor

# ------------ Definition of optimization models, risk models -------------

# In[1]:


"""
CVaR.py>
"""
import numpy as np
import math
from scipy.optimize import minimize
from scipy.stats.distributions import chi2
from numpy import linalg
from scipy.stats import gmean
from scipy.optimize import linprog
from scipy.stats import norm
from scipy import stats

# ----- function: CVaR model ------
"""
min     gamma + (1 / [(1 - alpha) * S]) * sum( z_s )
s.t.    z_s   >= 0,                 for s = 1, ..., S
        z_s   >= -r_s' x - gamma,   for s = 1, ..., S
        1' x  =  1,
        mu' x >= R
"""

def CVaR(returns, alpha):
    if returns.shape[1] == 1:
        x = np.asmatrix(1)
    else:
        mu = gmean(returns+1) - 1  # Estimate the geometric mean
        R = 1.1 * np.mean(mu)  # Set our target return
        t, p = returns.shape  # Determine the number of assets and scenarios
        # formulate the linear program
        # bound
        lst_t = np.ones(t)
        lst_p = np.ones(p)
        bounds = tuple((0, None) for i in lst_p) + tuple((0, None) for i in lst_t) +                  tuple((-math.inf, None) for i in np.ones(1))
        # inequality constraint matrix A and vector b
        A = np.vstack((np.hstack((-returns, -np.identity(t), -np.ones((t, 1)))),
                      np.hstack((-mu[np.newaxis], np.zeros((1, t)), np.zeros((1, 1))))))
        b = np.vstack((np.zeros((t, 1)), -R[np.newaxis]))
        # equality constraint matrix Aeq and vector beq
        Aeq = np.hstack((np.ones((1, p)), np.zeros((1, t)), np.zeros((1, 1))))
        beq = 1
        # coefficient of variables
        k = 1 / ((1 - alpha) * t)
        c = np.vstack((np.zeros((p, 1)), k*np.ones((t, 1)), np.ones((1, 1))))
        # linprog optimizer
        res = linprog(c, A_ub=A, b_ub=b, A_eq=Aeq, b_eq=beq, bounds=bounds)
        x = np.asmatrix(res.x[:p])
    
    return x


# In[2]:


"""
OLS.py>
"""

import numpy as np
from numpy.linalg import inv

# ----- function: OLS regression model ------
def OLS(returns, factRet):
    t, p = factRet.shape  # Number of observations and factors
    factRet = factRet.to_numpy(dtype=np.float64)
    returns = returns.to_numpy()
    X = np.hstack((np.ones((t, 1), dtype=np.float64), factRet))  # factor matrix;
    B = np.dot(np.dot(inv(np.dot(X.T, X)), X.T), returns)  # Regression coefficients;
    # separate B into alpha and betas
    a = B[0, :]
    V = B[1:, :]

    # residual variance
    ep = returns - np.dot(X, B)
    ep_2_sum = np.sum(ep ** 2, axis=0)  # sum each matrix column
    sigma_ep = 1 / (t - p - 1) * ep_2_sum
    D = np.diag(sigma_ep)

    # factor expected returns and covariance matrix
    f_bar = np.mean(factRet, axis=0)
    F = np.cov(factRet, rowvar=False)

    # Calc. asset expected returns and covariance matrices
    mu = a + np.dot(V.T, f_bar)
    Q = np.dot(np.dot(V.T, F), V) + D
    Q = (Q + Q.T) / 2

    # calc. adjusted R-squared
    r_bar = np.mean(returns, axis=0)
    RSS = ep_2_sum
    TSS = np.sum((returns - r_bar) ** 2, axis=0)
    R2 = 1 - RSS / TSS
    adj_R2 = 1 - (1 - R2) * ((t - 1) / (t - p - 1))
    adj_R2_bar = np.mean(adj_R2)

    return mu, Q, adj_R2_bar


# In[3]:


"""
RP.py>
"""

import numpy as np
from scipy.optimize import minimize

# ----- function: Risk Parity model ------

# risk budgeting optimization
def calculate_portfolio_var(x, Q):
    # function that calculates portfolio risk
    x = np.matrix(x)
    return (x * (Q * x.T))[0, 0]

def calculate_risk_contribution(x, Q):
    # function that calculates asset contribution to total risk
    x = np.matrix(x)
    sigma = np.sqrt(calculate_portfolio_var(x, Q))
    # Marginal Risk Contribution
    MRC = Q * x.T
    # Risk Contribution
    RC = np.multiply(MRC, x.T)/sigma
    return RC

def risk_budget_objective(x, pars):
    # calculate portfolio risk
    Q = pars[0]  # covariance table
    x_targ = pars[1]  # risk target in percent of portfolio risk
    sig_p = np.sqrt(calculate_portfolio_var(x, Q))  # portfolio sigma
    risk_target = np.asmatrix(np.multiply(sig_p, x_targ))
    asset_RC = calculate_risk_contribution(x, Q)
    J = sum(np.square(asset_RC-risk_target.T))[0, 0]  # sum of squared error
    return J

def total_weight_constraint(x):
    return np.sum(x)-1.0

def long_only_constraint(x):
    return x

def RP(Q):
    p = Q.shape[0]  # number of assets
    tol = 1e-14
    x_targ = np.full(p, 1 / p)  # define initial portfolio (equal risk)
    x0 = np.random.dirichlet(np.ones(p), size=1)
    cons = ({'type': 'eq', 'fun': total_weight_constraint},
            {'type': 'ineq', 'fun': long_only_constraint})
    res = minimize(risk_budget_objective, x0=x0, args=[Q, x_targ], method='SLSQP', constraints=cons,
                   tol=tol, options={'disp': False, 'maxiter': 10000})
    x = np.asmatrix(res.x)
    return x


# In[4]:


"""
robustMVO.py>
"""

import numpy as np
import math
from scipy.optimize import minimize
from scipy.stats.distributions import chi2
from numpy import linalg

# ----- function: robust MVO model ------

# model:
# min. lambda * (x' * Q * x) - mu' x + epsilon * norm (sqrtTh * x)
# s.t.  sum(x) == 1
#            x >= 0

def total_weight_constraint(x):
    return np.sum(x)-1.0

'''
def long_only_constraint(x):
    return x
'''

def sqrt_theta(Q, alpha):
    p = Q.shape[0]  # find the number of assets
    ep = math.sqrt(chi2.ppf(alpha, df=p))  # Define the radius of the uncertainty set
    theta = np.diag(np.diag(Q)) / p  # find the squared standard error of our expected returns
    sqrtTh = np.sqrt(theta)  # square root of theta
    return sqrtTh, ep

def objFn(x, pars):
    mu = pars[0]
    Q = pars[1]
    riskAver = pars[2]
    sqrtTh = pars[3]
    ep = pars[4]
    # objective function
    f = riskAver * np.dot(np.dot(x, Q), x.T) - np.dot(mu.T[np.newaxis], x.T) + ep * linalg.norm(sqrtTh * x)
    return f

def robustMVO(mu, Q, riskAver, alpha):
    p = Q.shape[0]  # find the number of assets
    sqrtTh, ep = sqrt_theta(Q, alpha)
    x0 = np.random.dirichlet(np.ones(p), size=1)[0] # initial weight
    tol = 1e-14
    cons = ({'type': 'eq', 'fun': total_weight_constraint})
    lst = np.ones(p)
    bounds = tuple((0, 1) for i in lst)
    res = minimize(objFn, x0=x0, args=[mu, Q, riskAver, sqrtTh, ep], method='SLSQP', constraints=cons,
                   bounds=bounds, tol=tol, options={'disp': False, 'maxiter': 10000})
    x = np.asmatrix(res.x)
    return x


# In[ ]:


def VaRCalc(returns):
    # VaR and CVaR Calculation
    
    period = 1. # horizon of 1 day
    mu_p = np.mean(returns)/30
    sig_p = np.std(returns)*np.sqrt(period/30)
    alpha = 0.01

    VaR_1 = norm.ppf(1-alpha)*sig_p - mu_p 
    value_1 =portfValue_wocf[-1]*VaR_1
    print("99% 1 day VaR :", round(VaR_1*100,2),"% or", round(value_1),"$")

    CVaR_1 = alpha**-1 * norm.pdf(norm.ppf(alpha))*sig_p - mu_p
    value_C_1 =portfValue_wocf[-1]*CVaR_1
    print("99% 1 day CVaR/ES :", round(CVaR_1*100,2),"% or", round(value_C_1),"$")

    period = 10. # horizon of 10 days
    mu_p = np.mean(returns)/30*period
    sig_p = np.std(returns) * np.sqrt(period/30)
    VaR_2 = norm.ppf(1-alpha)*sig_p - mu_p
    value_2=portfValue_wocf[-1]*VaR_2
    print("99% 10 day VaR :", round(VaR_2*100,2),"% or", round(value_2),"$")

    CVaR_2 = alpha**-1 * norm.pdf(norm.ppf(alpha))*sig_p - mu_p
    value_C_2 =portfValue_wocf[-1]*CVaR_2
    print("99% 10 day CVaR/ES :", round(CVaR_2*100,2),"% or", round(value_C_2),"$")
    
    return VaR_1, value_1, CVaR_1, value_C_1, VaR_2, value_2, CVaR_2, value_C_2 


# In[4]:


def my_rolling_sharpe(y):
    return np.sqrt(6) * (y.mean() / y.std())

def my_rolling_vol(y):
    return np.sqrt(6) * (y.std())


# In[5]:


# import packages

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook
import math
from datetime import datetime, timedelta
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from copulas.multivariate import VineCopula
import statsmodels.api as sm
import seaborn as sns
pd.set_option('display.max_columns', 20)


# ------------------Backtest-----------------------
# 
# The following two blocks of code are designed specifically for backtesting. Some inputs such as startDate and reportDate are manually set to be 2014-03-01 and 2016-03-01 respectively to fix the backtesting period. Backtest period should be before the investment period starts. We use our backtest results to compare the three asset allocation models' performance and select the optimal model.
# 
# IMPORTANT: when running the production model, need to comment out the two blocks of code for backtesting 

# Backtest Period from 2014.3.1 to 2016.3.1
# rebalancing period=24 month

# In[6]:


# backtest block 1

dir = os.getcwd()  # wd

# 1. Link to the Robo Advisor UI and get user inputs
wb = load_workbook(dir + r'\AWGP_RoboAdvisor.xlsx')
ws = wb['UI']
name = ws[list(wb.defined_names['name'].destinations)[0][1]].value
age = ws[list(wb.defined_names['age'].destinations)[0][1]].value  # client age
cap = ws[list(wb.defined_names['cap'].destinations)[0][1]].value  # initial capital
startDate = ws[list(wb.defined_names['startDate'].destinations)[0][1]].value
injection = ws[list(wb.defined_names['injection'].destinations)[0][1]].value  # periodic injection amount
injectFreq = ws[list(wb.defined_names['injectFreq'].destinations)[0][1]].value
ESG = ws[list(wb.defined_names['ESG'].destinations)[0][1]].value  # whether to include ESG or not
ESGperc = ws[list(wb.defined_names['ESGperc'].destinations)[0][1]].value  # ESG percentage
caliYr = ws[list(wb.defined_names['caliYr'].destinations)[0][1]].value  # calibration period length
leverage = ws[list(wb.defined_names['leverage'].destinations)[0][1]].value  # percentage of leverage
reportDate = ws[list(wb.defined_names['reportDate'].destinations)[0][1]].value
rebalFreq = ws[list(wb.defined_names['rebalFreq'].destinations)[0][1]].value  # rebalancing frequency is every 6 months
cost = ws[list(wb.defined_names['cost'].destinations)[0][1]].value  # transaction cost per unit of ETF traded
countryPerc = ws[list(wb.defined_names['countryPerc'].destinations)[0][1]].value  # USD:CAD exposure. 50% means 50/50
equityFI_perc = ws[list(wb.defined_names['equityFI_perc'].destinations)[0][1]].value  # equity+FI percentage in portf
alpha = ws[list(wb.defined_names['alpha'].destinations)[0][1]].value  # confidence level
riskAver = ws[list(wb.defined_names['riskAver'].destinations)[0][1]].value  # risk tolerance


# Replace Parameters
startDate=datetime(2014, 3, 1)
reportDate=datetime(2016, 3, 1)
caliYr=8
rebalFreq=24 # prevent rebalancing in backtesting
injectFreq = 100 # don't inject


# 2. Link to the investment pool (ESG + non_ESG ETFs) and factors
# data is from 2007/01/01

# ETFs
df_ESG = pd.read_excel(dir + r'\ESG_Prices.xlsx', sheet_name='Prices', index_col=0, engine='openpyxl')
df_nonESG = pd.read_excel(dir + r'\NonESG_Prices.xlsx', sheet_name='Prices', index_col=0, engine='openpyxl')
df_nonESG = df_nonESG.loc[:, ~df_nonESG.columns.str.contains('^Unnamed')]
ret_ESG = df_ESG.pct_change().iloc[1:].loc[:reportDate]
ret_nonESG = df_nonESG.pct_change().iloc[1:].loc[:reportDate]

# select ETFs with low correlations
def select_ETF(df1, df2):
    df = df1.copy()
    cor = df.corr().abs()
    # Select upper/lower triangle of correlation matrix
    upper_triangle = cor.where(np.triu(np.ones(cor.shape), k=1).astype(bool))
    # Find features with correlation greater than 0.85
    drop_lst1 = [column for column in upper_triangle.columns if any(upper_triangle[column] > 0.8)]
    print('number of features dropped that have high correlations: ' + str(len(drop_lst1)))
    # Drop features
    df2.drop(drop_lst1, axis=1, inplace=True)
    print('number of features left: ' + str(len(df2.columns)))
    return df2

ret_ESG = select_ETF(ret_ESG.loc[:startDate], ret_ESG)
ret_nonESG = select_ETF(ret_nonESG.loc[:startDate], ret_nonESG)

# dates and tickers
dates = ret_ESG.index  # all dates of our dataset
tickers_nonESG = list(ret_nonESG.columns)
tickers_ESG = list(ret_ESG.columns)
df_ESG = pd.DataFrame(df_ESG, columns=tickers_ESG)  # ESG dataframe
df_nonESG = pd.DataFrame(df_nonESG, columns=tickers_nonESG)  # non-ESG dataframe
print("nonESG ETFs left:" + str(tickers_nonESG))
print("ESG ETFs left:" + str(tickers_ESG))


# ----- input: final selected ETFs - need to update the list manually according to the above print output ------
allocation_model = 'RP'
tickers_nonESG_equity_FI_USD = ['XNTK', 'KCE', 'XLE', 'XLV', 'XLU', 'SHY', 'TLT', 'TIP', 'LQD', 'MHWIX']
tickers_nonESG_other_USD = ['GLD', 'GSG']
tickers_nonESG_equity_FI_CAD = ['XIT.TO', 'XBB.TO', 'XRB.TO']
tickers_nonESG_other_CAD = ['XRE.TO']
tickers_ESG_equity_USD = ['XLB']
# --------------------------------------------------------------------------------------------------------------


# In[7]:


# backtest block 2

# factors
factors = pd.read_csv(dir + r'\FF_5_Factor.csv', index_col=0,parse_dates=True)
factors = factors.loc[dates]
rf = pd.DataFrame(factors['RF'], columns=['RF'])  # risk-free rate data
ret_factor = factors.drop('RF', axis=1)  # factor returns data
rf.set_index(dates, inplace=True)
ret_factor.set_index(dates, inplace=True)

# CAD/USD FX rate
FX = pd.read_csv(dir + r'\USD_CAD_Historical_Data.csv', index_col=0,parse_dates=True)  # FX rate
FX = FX.loc[dates]

# Excess Return
exRet_nonESG = ret_nonESG.sub(rf['RF'], 0)  # non ESG ETFs excess return dataset
exRet_ESG = ret_ESG.sub(rf['RF'], 0)

# leverage
init_cap = cap * (1+leverage)  # total initial capital including leverage  # initial capital + leverage
loan = cap * leverage  # loan at start of investment period; interest rate = RFR

# 3. system implementation
exRet_portf = exRet_nonESG.copy()  # portfolio excess return dataset for non ESG ETFs
# separate by asset class
exRet_equity_FI_USD = pd.DataFrame(exRet_portf, columns=tickers_nonESG_equity_FI_USD)
exRet_other_USD = pd.DataFrame(exRet_portf, columns=tickers_nonESG_other_USD)
exRet_equity_FI_CAD = pd.DataFrame(exRet_portf, columns=tickers_nonESG_equity_FI_CAD)
exRet_other_CAD = pd.DataFrame(exRet_portf, columns=tickers_nonESG_other_CAD)

if ESG == 'No':
    price_portf = df_nonESG.iloc[1:].loc[:reportDate].copy()  # investment portfolio's monthly price dataset for non ESG
    # apply FX on USD non ESG ETFs price
    price_portf[tickers_nonESG_equity_FI_USD + tickers_nonESG_other_USD] =         price_portf[tickers_nonESG_equity_FI_USD + tickers_nonESG_other_USD].mul(FX['Price'], axis=0)
    price_portf = pd.concat([price_portf[tickers_nonESG_equity_FI_USD], price_portf[tickers_nonESG_other_USD],
                             price_portf[tickers_nonESG_equity_FI_CAD], price_portf[tickers_nonESG_other_CAD]],
                            axis=1).copy()

if ESG == 'Yes':
    price_portf = pd.concat([df_nonESG, df_ESG], axis=1).iloc[1:].loc[:reportDate].copy()  # non ESG+ESG ETFs price data
    exRet_portf = pd.concat([exRet_nonESG, exRet_ESG], axis=1).copy()  # non ESG+ESG ETFs excess return data
    # ESG ETFs returns
    exRet_ESG_equity_FI_USD = pd.DataFrame(exRet_portf, columns=tickers_ESG_equity_USD)
    # apply FX on USD ETFs for both non ESG and ESG
    price_portf[tickers_nonESG_equity_FI_USD + tickers_nonESG_other_USD + tickers_ESG_equity_USD] =         price_portf[tickers_nonESG_equity_FI_USD + tickers_nonESG_other_USD + tickers_ESG_equity_USD].mul(FX['Price'],
                                                                                                          axis=0)
    price_portf = pd.concat([price_portf[tickers_nonESG_equity_FI_USD], price_portf[tickers_nonESG_other_USD],
                             price_portf[tickers_nonESG_equity_FI_CAD], price_portf[tickers_nonESG_other_CAD],
                             price_portf[tickers_ESG_equity_USD]],
                            axis=1).copy()


n = exRet_portf.shape[1]  # number of ETFs in the portfolio
n_equity_FI_USD = exRet_equity_FI_USD.shape[1]  # number of USD ETFs (equity+FI)
n_other_USD = exRet_other_USD.shape[1]  # number of USD ETFs (equity+FI)
n_equity_FI_CAD = exRet_equity_FI_CAD.shape[1]  # number of CAD ETFs (equity+FI)
n_other_CAD = exRet_other_CAD.shape[1]  # number of CAD ETFs (equity+FI)
n_ESG_equity_FI_USD = exRet_ESG_equity_FI_USD.shape[1]  # number of ESG USD ETFs (equity+FI)
testStart = exRet_portf.loc[startDate:].index[0]  # each rebalancing period start date
# testEnd = exRet_portf.loc[startDate:].index[rebalFreq] - timedelta(days=1)   # each rebalancing period end date
testEnd=reportDate
caliEnd = testStart - timedelta(days=1)  # the end of each calibration period
NoPeriods = math.ceil((len(exRet_portf.loc[startDate:]) - 1) / rebalFreq)  # number of rebalancing periods
testMonths = len(exRet_portf.loc[startDate:])  # number of months in the test period
testDate = exRet_portf.loc[startDate:].index  # all dates in the test period


# Preallocate space
x_equity_FI_USD = np.zeros((n_equity_FI_USD, NoPeriods))  # USD equity + FI non ESG ETFs weight
x_other_USD = np.zeros((n_other_USD, NoPeriods))  # USD REIT + commodity non ESG ETFs weight
x_equity_FI_CAD = np.zeros((n_equity_FI_CAD, NoPeriods))  # CAD equity + FI non ESG ETFs weight
x_other_CAD = np.zeros((n_other_CAD, NoPeriods))  # CAD REIT + commodity non ESG ETFs weight
x_ESG_equity_FI_USD = np.zeros((n_ESG_equity_FI_USD, NoPeriods))  # USD ESG equity ETFs weight
x = np.zeros((n, NoPeriods))  # all ETFs weights
x0 = np.zeros((n, NoPeriods))  # all ETFs weights before rebalancing
adj_R2_bar = np.zeros(NoPeriods)
NoShares = np.zeros(n)  # all ETFs number of shares before rebalancing
NoShares_new = np.zeros(n)  # all ETFs number of shares after rebalancing
currentVal = np.zeros(NoPeriods)  # last observed portfolio value during current calibration period
portfValue = np.zeros(testMonths)  # portfolio value for each month in the test period
turnover = np.zeros(NoPeriods)
testDate2 = np.zeros(NoPeriods, dtype='datetime64[ns]')  # last day during each rebalancing period

# loop through each rebalancing period
for t in range(NoPeriods):
    # excess return of portfolio for the calibration period
    ret_cali_equity_FI_USD = exRet_equity_FI_USD.loc[caliEnd-timedelta(days=caliYr*365.25):caliEnd]
    ret_cali_other_USD = exRet_other_USD.loc[caliEnd - timedelta(days=caliYr * 365.25):caliEnd]
    ret_cali_equity_FI_CAD = exRet_equity_FI_CAD.loc[caliEnd-timedelta(days=caliYr*365.25):caliEnd]
    ret_cali_other_CAD = exRet_other_CAD.loc[caliEnd - timedelta(days=caliYr * 365.25):caliEnd]
    if ESG == 'Yes':
        ret_cali_ESG_equity_FI_USD = exRet_ESG_equity_FI_USD.loc[caliEnd - timedelta(days=caliYr * 365.25):caliEnd]

    # factors return for the calibration period
    ret_factor_cali = ret_factor.loc[caliEnd-timedelta(days=caliYr*365.25):caliEnd]

    currPrice = price_portf.loc[:caliEnd].iloc[-1]  # last observed price during calibration period
    testPrice = price_portf.loc[testStart:testEnd]  # rebalancing period prices
    testPeriod = len(price_portf.loc[testStart:testEnd])
    testDate2[t] = price_portf.loc[testStart:testEnd].index[-1].value  # last day during each rebalancing period

    if t == 0:
        # last observed portfolio value during current calibration period
        currentVal[t] = init_cap
    else:
        currentVal[t] = np.dot(currPrice, NoShares) + injection - trans_cost

        # store the current asset weights (before rebalance)
        x0[:, t] = (currPrice * NoShares) / currentVal[t]

    # portfolio construction
    # factor model: OLS calibration + optimization model, apply on each asset class
    # Risk-Parity
    if allocation_model == 'RP':
        if ESG == 'No':
            mu1, Q1, adj_R2_bar1 = OLS(ret_cali_equity_FI_USD, ret_factor_cali)
            x_equity_FI_USD[:, t] = RP(Q1) * countryPerc * equityFI_perc

            mu2, Q2, adj_R2_bar2 = OLS(ret_cali_other_USD, ret_factor_cali)
            x_other_USD[:, t] = RP(Q2) * countryPerc * (1-equityFI_perc)

            mu3, Q3, adj_R2_bar3 = OLS(ret_cali_equity_FI_CAD, ret_factor_cali)
            x_equity_FI_CAD[:, t] = RP(Q3) * countryPerc * equityFI_perc

            mu4, Q4, adj_R2_bar4 = OLS(ret_cali_other_CAD, ret_factor_cali)
            x_other_CAD[:, t] = RP(Q4) * countryPerc * (1-equityFI_perc)

        elif ESG == 'Yes':
            mu1, Q1, adj_R2_bar1 = OLS(ret_cali_equity_FI_USD, ret_factor_cali)
            x_equity_FI_USD[:, t] = RP(Q1) * (countryPerc * equityFI_perc - ESGperc)  # 20%

            mu2, Q2, adj_R2_bar2 = OLS(ret_cali_other_USD, ret_factor_cali)
            x_other_USD[:, t] = RP(Q2) * countryPerc * (1 - equityFI_perc)  # 10%

            mu3, Q3, adj_R2_bar3 = OLS(ret_cali_equity_FI_CAD, ret_factor_cali)
            x_equity_FI_CAD[:, t] = RP(Q3) * countryPerc * equityFI_perc  # 40%

            mu4, Q4, adj_R2_bar4 = OLS(ret_cali_other_CAD, ret_factor_cali)
            x_other_CAD[:, t] = RP(Q4) * countryPerc * (1 - equityFI_perc)  # 10%

            mu5, Q5, adj_R2_bar5 = OLS(ret_cali_ESG_equity_FI_USD, ret_factor_cali)
            x_ESG_equity_FI_USD[:, t] = RP(Q5) * ESGperc  # 20%
    # CVaR
    elif allocation_model == 'CVaR':
        if ESG == 'No':
            mu1, Q1, adj_R2_bar1 = OLS(ret_cali_equity_FI_USD, ret_factor_cali)
            x_equity_FI_USD[:, t] = CVaR(ret_cali_equity_FI_USD, alpha) * countryPerc * equityFI_perc

            mu2, Q2, adj_R2_bar2 = OLS(ret_cali_other_USD, ret_factor_cali)
            x_other_USD[:, t] = CVaR(ret_cali_other_USD, alpha) * countryPerc * (1 - equityFI_perc)

            mu3, Q3, adj_R2_bar3 = OLS(ret_cali_equity_FI_CAD, ret_factor_cali)
            x_equity_FI_CAD[:, t] = CVaR(ret_cali_equity_FI_CAD, alpha) * countryPerc * equityFI_perc

            mu4, Q4, adj_R2_bar4 = OLS(ret_cali_other_CAD, ret_factor_cali)
            x_other_CAD[:, t] = CVaR(ret_cali_other_CAD, alpha) * countryPerc * (1 - equityFI_perc)

        elif ESG == 'Yes':
            mu1, Q1, adj_R2_bar1 = OLS(ret_cali_equity_FI_USD, ret_factor_cali)
            x_equity_FI_USD[:, t] = CVaR(ret_cali_equity_FI_USD, alpha) * (countryPerc * equityFI_perc - ESGperc)  # 20%

            mu2, Q2, adj_R2_bar2 = OLS(ret_cali_other_USD, ret_factor_cali)
            x_other_USD[:, t] = CVaR(ret_cali_other_USD, alpha) * countryPerc * (1 - equityFI_perc)  # 10%

            mu3, Q3, adj_R2_bar3 = OLS(ret_cali_equity_FI_CAD, ret_factor_cali)
            x_equity_FI_CAD[:, t] = CVaR(ret_cali_equity_FI_CAD, alpha) * countryPerc * equityFI_perc  # 40%

            mu4, Q4, adj_R2_bar4 = OLS(ret_cali_other_CAD, ret_factor_cali)
            x_other_CAD[:, t] = CVaR(ret_cali_other_CAD, alpha) * countryPerc * (1 - equityFI_perc)  # 10%

            mu5, Q5, adj_R2_bar5 = OLS(ret_cali_ESG_equity_FI_USD, ret_factor_cali)
            x_ESG_equity_FI_USD[:, t] = CVaR(ret_cali_ESG_equity_FI_USD, alpha) * ESGperc  # 20%
    # Robust MVO
    elif allocation_model == 'robustMVO':
        if ESG == 'No':
            mu1, Q1, adj_R2_bar1 = OLS(ret_cali_equity_FI_USD, ret_factor_cali)
            x_equity_FI_USD[:, t] = robustMVO(mu1, Q1, riskAver, alpha) * countryPerc * equityFI_perc

            mu2, Q2, adj_R2_bar2 = OLS(ret_cali_other_USD, ret_factor_cali)
            x_other_USD[:, t] = robustMVO(mu2, Q2, riskAver, alpha) * countryPerc * (1 - equityFI_perc)

            mu3, Q3, adj_R2_bar3 = OLS(ret_cali_equity_FI_CAD, ret_factor_cali)
            x_equity_FI_CAD[:, t] = robustMVO(mu3, Q3, riskAver, alpha) * countryPerc * equityFI_perc

            mu4, Q4, adj_R2_bar4 = OLS(ret_cali_other_CAD, ret_factor_cali)
            x_other_CAD[:, t] = robustMVO(mu4, Q4, riskAver, alpha) * countryPerc * (1 - equityFI_perc)

        elif ESG == 'Yes':
            mu1, Q1, adj_R2_bar1 = OLS(ret_cali_equity_FI_USD, ret_factor_cali)
            x_equity_FI_USD[:, t] = robustMVO(mu1, Q1, riskAver, alpha) * (countryPerc * equityFI_perc - ESGperc)

            mu2, Q2, adj_R2_bar2 = OLS(ret_cali_other_USD, ret_factor_cali)
            x_other_USD[:, t] = robustMVO(mu2, Q2, riskAver, alpha) * countryPerc * (1 - equityFI_perc)

            mu3, Q3, adj_R2_bar3 = OLS(ret_cali_equity_FI_CAD, ret_factor_cali)
            x_equity_FI_CAD[:, t] = robustMVO(mu3, Q3, riskAver, alpha) * countryPerc * equityFI_perc

            mu4, Q4, adj_R2_bar4 = OLS(ret_cali_other_CAD, ret_factor_cali)
            x_other_CAD[:, t] = robustMVO(mu4, Q4, riskAver, alpha) * countryPerc * (1 - equityFI_perc)

            mu5, Q5, adj_R2_bar5 = OLS(ret_cali_ESG_equity_FI_USD, ret_factor_cali)
            x_ESG_equity_FI_USD[:, t] = robustMVO(mu5, Q5, riskAver, alpha) * ESGperc
    # adjusted R-square: check model fitting
    if ESG == 'No':
        adj_R2_bar[t] = (adj_R2_bar1 + adj_R2_bar2 + adj_R2_bar3 + adj_R2_bar4) / 4
        # new asset weight
        x[:, t] = np.append(np.append(np.append(x_equity_FI_USD[:, t], x_other_USD[:, t]), x_equity_FI_CAD[:, t]),
                            x_other_CAD[:, t])
        tickers = tickers_nonESG_equity_FI_USD + tickers_nonESG_other_USD + tickers_nonESG_equity_FI_CAD +                   tickers_nonESG_other_CAD
        # weight dataframe
        weights = pd.DataFrame(x.T, columns=tickers)
    elif ESG == 'Yes':
        adj_R2_bar[t] = (adj_R2_bar1 + adj_R2_bar2 + adj_R2_bar3 + adj_R2_bar4 + adj_R2_bar5) / 5
        x[:, t] = np.append(np.append(np.append(np.append(x_equity_FI_USD[:, t], x_other_USD[:, t]),
                                                x_equity_FI_CAD[:, t]), x_other_CAD[:, t]), x_ESG_equity_FI_USD[:, t])
        tickers = tickers_nonESG_equity_FI_USD + tickers_nonESG_other_USD + tickers_nonESG_equity_FI_CAD +                   tickers_nonESG_other_CAD + tickers_ESG_equity_USD
        weights = pd.DataFrame(x.T, columns=tickers)

    NoShares_new = x[:, t] * currentVal[t] / currPrice  # Number of shares to be allocated to each ETFs
    trans_cost = sum(abs(NoShares_new - NoShares) * cost)
    NoShares = NoShares_new

    # calculate the  portfolio value at each month during the test period
    for j in range(testPeriod):
        index_portf = np.where(exRet_portf.loc[startDate:].index == testStart)[0][0]
        # loan interest accumulation
        rf_7yr = 1.56  # used 7-year treasury rate from 2014-03-01 which is 1.56%
        loan_pay = loan * (1 + rf_7yr / 100) ** ((index_portf+j)/12)  # accumulated by 7-year treasury rate

        if j == 0:
            price_portf.loc[testStart]
            portfValue[index_portf+j] = np.dot(testPrice, NoShares_new)[j] - loan_pay - trans_cost
        else:
            portfValue[index_portf+j] = np.dot(testPrice, NoShares_new)[j] - loan_pay

    # calculate the turnover rate
    if t > 0:
        turnover[t] = sum(abs(x[:, t] - x0[:, t]))

    # Update the calibration and out-of-sample test periods
    if t < 9:
        testStart = exRet_portf.loc[testStart:].index[6]  # each rebalancing period start date
        testEnd = exRet_portf.loc[testStart:].index[6] - timedelta(days=1)  # each rebalancing period end date
        caliEnd = testStart - timedelta(days=1)  # the end of each calibration period
    elif t == 9:
        testStart = exRet_portf.loc[testStart:].index[6]  # each rebalancing period start date
        testEnd = exRet_portf.loc[testStart:].index[-1]  # each rebalancing period end date
        caliEnd = testStart - timedelta(days=1)  # the end of each calibration period


# outputs:
print('portfolio value evolution:' + str(portfValue.T))
print(weights)  # portfolio weighting
# print('portfolio weighting:' + str(x.T))
print('turnover:' + str(turnover))
print('adjusted R-squared:' + str(adj_R2_bar))
print('date:' + str(testDate))

# Check the correlation heatmap for our dataset
fig, ax = plt.subplots(figsize=(10,10))
corr = exRet_portf.corr()
sns.heatmap(corr, xticklabels=corr.columns, yticklabels=corr.columns, linewidths=.1, vmin=-1, vmax=1)
ax.set_title('Heatmap for all ETFs excess returns', fontsize=20)
plt.show()


# -------------------Backtest END---------------------------



# ------------- main system implementation ---------------
# 
# The following two blocks of code are for our main system implementation that grabs real clients inputs from the user interface "AWGP_RoboAdvisor.xlsx". When executing the main program below, the above two blocks of backtesting code should not be run (comment out).

# In[8]:


# ================== Robo Advisor <All Weather Global Portfolio> ==============================
dir = os.getcwd()  # wd

# 1. Link to the Robo Advisor UI and get user inputs
wb = load_workbook(dir + r'\AWGP_RoboAdvisor.xlsx')
ws = wb['UI']
name = ws[list(wb.defined_names['name'].destinations)[0][1]].value
age = ws[list(wb.defined_names['age'].destinations)[0][1]].value  # client age
cap = ws[list(wb.defined_names['cap'].destinations)[0][1]].value  # initial capital
startDate = ws[list(wb.defined_names['startDate'].destinations)[0][1]].value
injection = ws[list(wb.defined_names['injection'].destinations)[0][1]].value  # periodic injection amount
injectFreq = ws[list(wb.defined_names['injectFreq'].destinations)[0][1]].value
ESG = ws[list(wb.defined_names['ESG'].destinations)[0][1]].value  # whether to include ESG or not
ESGperc = ws[list(wb.defined_names['ESGperc'].destinations)[0][1]].value  # ESG percentage
caliYr = ws[list(wb.defined_names['caliYr'].destinations)[0][1]].value  # calibration period length
leverage = ws[list(wb.defined_names['leverage'].destinations)[0][1]].value  # percentage of leverage
reportDate = ws[list(wb.defined_names['reportDate'].destinations)[0][1]].value
rebalFreq = ws[list(wb.defined_names['rebalFreq'].destinations)[0][1]].value  # rebalancing frequency is every 6 months
cost = ws[list(wb.defined_names['cost'].destinations)[0][1]].value  # transaction cost per unit of ETF traded
countryPerc = ws[list(wb.defined_names['countryPerc'].destinations)[0][1]].value  # USD:CAD exposure. 50% means 50/50
equityFI_perc = ws[list(wb.defined_names['equityFI_perc'].destinations)[0][1]].value  # equity+FI percentage in portf
alpha = ws[list(wb.defined_names['alpha'].destinations)[0][1]].value  # confidence level
riskAver = ws[list(wb.defined_names['riskAver'].destinations)[0][1]].value  # risk tolerance

# 2. Link to the investment pool (ESG + non_ESG ETFs) and factors
# data is from 2007/01/01

# ETFs
df_ESG = pd.read_excel(dir + r'\ESG_Prices.xlsx', sheet_name='Prices', index_col=0, engine='openpyxl')
df_nonESG = pd.read_excel(dir + r'\NonESG_Prices.xlsx', sheet_name='Prices', index_col=0, engine='openpyxl')
df_nonESG = df_nonESG.loc[:, ~df_nonESG.columns.str.contains('^Unnamed')]
ret_ESG = df_ESG.pct_change().iloc[1:].loc[:reportDate]
ret_nonESG = df_nonESG.pct_change().iloc[1:].loc[:reportDate]

# select ETFs with low correlations
def select_ETF(df1, df2):
    df = df1.copy()
    cor = df.corr().abs()
    # Select upper/lower triangle of correlation matrix
    upper_triangle = cor.where(np.triu(np.ones(cor.shape), k=1).astype(bool))
    # Find features with correlation greater than 0.85
    drop_lst1 = [column for column in upper_triangle.columns if any(upper_triangle[column] > 0.8)]
    print('number of features dropped that have high correlations: ' + str(len(drop_lst1)))
    # Drop features
    df2.drop(drop_lst1, axis=1, inplace=True)
    print('number of features left: ' + str(len(df2.columns)))
    return df2

ret_ESG = select_ETF(ret_ESG.loc[:startDate], ret_ESG)
ret_nonESG = select_ETF(ret_nonESG.loc[:startDate], ret_nonESG)

# dates and tickers
dates = ret_ESG.index  # all dates of our dataset
tickers_nonESG = list(ret_nonESG.columns)
tickers_ESG = list(ret_ESG.columns)
df_ESG = pd.DataFrame(df_ESG, columns=tickers_ESG)  # ESG dataframe
df_nonESG = pd.DataFrame(df_nonESG, columns=tickers_nonESG)  # non-ESG dataframe
print("nonESG ETFs left:" + str(tickers_nonESG))
print("ESG ETFs left:" + str(tickers_ESG))


# ----- input: final selected ETFs - need to update the list manually according to the above print output ------
allocation_model = 'CVaR'
tickers_nonESG_equity_FI_USD = ['XNTK', 'KCE', 'XLE', 'XLV', 'XLU', 'SHY', 'TLT', 'TIP', 'LQD', 'MHWIX']
tickers_nonESG_other_USD = ['IYR', 'GLD', 'GSG']
tickers_nonESG_equity_FI_CAD = ['XIT.TO', 'XBB.TO', 'XRB.TO']
tickers_nonESG_other_CAD = ['XRE.TO']
tickers_ESG_equity_USD = ['XLB', 'SOXX']
# --------------------------------------------------------------------------------------------------------------


# In[9]:


# factors
factors = pd.read_csv(dir + r'\FF_5_Factor.csv', index_col=0)
rf = pd.DataFrame(factors['RF'].iloc[1:], columns=['RF'])  # risk-free rate data
ret_factor = factors.drop('RF', axis=1).iloc[1:]  # factor returns data
rf.set_index(dates, inplace=True)
ret_factor.set_index(dates, inplace=True)

# CAD/USD FX rate
FX = pd.read_csv(dir + r'\USD_CAD_Historical_Data.csv', index_col=0)  # FX rate
FX.set_index(dates, inplace=True)

# Excess Return
exRet_nonESG = ret_nonESG.sub(rf['RF'], 0)  # non ESG ETFs excess return dataset
exRet_ESG = ret_ESG.sub(rf['RF'], 0)

# leverage
init_cap = cap * (1+leverage)  # total initial capital including leverage  # initial capital + leverage
loan = cap * leverage  # loan at start of investment period; interest rate = RFR

# 3. system implementation
exRet_portf = exRet_nonESG.copy()  # portfolio excess return dataset for non ESG ETFs
# separate by asset class
exRet_equity_FI_USD = pd.DataFrame(exRet_portf, columns=tickers_nonESG_equity_FI_USD)
exRet_other_USD = pd.DataFrame(exRet_portf, columns=tickers_nonESG_other_USD)
exRet_equity_FI_CAD = pd.DataFrame(exRet_portf, columns=tickers_nonESG_equity_FI_CAD)
exRet_other_CAD = pd.DataFrame(exRet_portf, columns=tickers_nonESG_other_CAD)

if ESG == 'No':
    price_portf = df_nonESG.iloc[1:].loc[:reportDate].copy()  # investment portfolio's monthly price dataset for non ESG
    # apply FX on USD non ESG ETFs price
    price_portf[tickers_nonESG_equity_FI_USD + tickers_nonESG_other_USD] =         price_portf[tickers_nonESG_equity_FI_USD + tickers_nonESG_other_USD].mul(FX['Price'], axis=0)
    price_portf = pd.concat([price_portf[tickers_nonESG_equity_FI_USD], price_portf[tickers_nonESG_other_USD],
                             price_portf[tickers_nonESG_equity_FI_CAD], price_portf[tickers_nonESG_other_CAD]],
                            axis=1).copy()

if ESG == 'Yes':
    price_portf = pd.concat([df_nonESG, df_ESG], axis=1).iloc[1:].loc[:reportDate].copy()  # non ESG+ESG ETFs price data
    exRet_portf = pd.concat([exRet_nonESG, exRet_ESG], axis=1).copy()  # non ESG+ESG ETFs excess return data
    # ESG ETFs returns
    exRet_ESG_equity_FI_USD = pd.DataFrame(exRet_portf, columns=tickers_ESG_equity_USD)
    # apply FX on USD ETFs for both non ESG and ESG
    price_portf[tickers_nonESG_equity_FI_USD + tickers_nonESG_other_USD + tickers_ESG_equity_USD] =         price_portf[tickers_nonESG_equity_FI_USD + tickers_nonESG_other_USD + tickers_ESG_equity_USD].mul(FX['Price'],
                                                                                                          axis=0)
    price_portf = pd.concat([price_portf[tickers_nonESG_equity_FI_USD], price_portf[tickers_nonESG_other_USD],
                             price_portf[tickers_nonESG_equity_FI_CAD], price_portf[tickers_nonESG_other_CAD],
                             price_portf[tickers_ESG_equity_USD]],
                            axis=1).copy()


n = exRet_portf.shape[1]  # number of ETFs in the portfolio
n_equity_FI_USD = exRet_equity_FI_USD.shape[1]  # number of USD ETFs (equity+FI)
n_other_USD = exRet_other_USD.shape[1]  # number of USD ETFs (equity+FI)
n_equity_FI_CAD = exRet_equity_FI_CAD.shape[1]  # number of CAD ETFs (equity+FI)
n_other_CAD = exRet_other_CAD.shape[1]  # number of CAD ETFs (equity+FI)
n_ESG_equity_FI_USD = exRet_ESG_equity_FI_USD.shape[1]  # number of ESG USD ETFs (equity+FI)
testStart = exRet_portf.loc[startDate:].index[0]  # each rebalancing period start date
testEnd = exRet_portf.loc[startDate:].index[6] - timedelta(days=1)   # each rebalancing period end date
caliEnd = testStart - timedelta(days=1)  # the end of each calibration period
NoPeriods = math.ceil((len(exRet_portf.loc[startDate:]) - 1) / rebalFreq)  # number of rebalancing periods
testMonths = len(exRet_portf.loc[startDate:])  # number of months in the test period
testDate = exRet_portf.loc[startDate:].index  # all dates in the test period


# Preallocate space
x_equity_FI_USD = np.zeros((n_equity_FI_USD, NoPeriods))  # USD equity + FI non ESG ETFs weight
x_other_USD = np.zeros((n_other_USD, NoPeriods))  # USD REIT + commodity non ESG ETFs weight
x_equity_FI_CAD = np.zeros((n_equity_FI_CAD, NoPeriods))  # CAD equity + FI non ESG ETFs weight
x_other_CAD = np.zeros((n_other_CAD, NoPeriods))  # CAD REIT + commodity non ESG ETFs weight
x_ESG_equity_FI_USD = np.zeros((n_ESG_equity_FI_USD, NoPeriods))  # USD ESG equity ETFs weight
x = np.zeros((n, NoPeriods))  # all ETFs weights
x0 = np.zeros((n, NoPeriods))  # all ETFs weights before rebalancing
adj_R2_bar = np.zeros(NoPeriods)
NoShares = np.zeros(n)  # all ETFs number of shares before rebalancing
NoShares_new = np.zeros(n)  # all ETFs number of shares after rebalancing
currentVal = np.zeros(NoPeriods)  # last observed portfolio value during current calibration period
portfValue = np.zeros(testMonths)  # portfolio value for each month in the test period
turnover = np.zeros(NoPeriods)
testDate2 = np.zeros(NoPeriods, dtype='datetime64[ns]')  # last day during each rebalancing period

# loop through each rebalancing period
for t in range(NoPeriods):
    # excess return of portfolio for the calibration period
    ret_cali_equity_FI_USD = exRet_equity_FI_USD.loc[caliEnd-timedelta(days=caliYr*365.25):caliEnd]
    ret_cali_other_USD = exRet_other_USD.loc[caliEnd - timedelta(days=caliYr * 365.25):caliEnd]
    ret_cali_equity_FI_CAD = exRet_equity_FI_CAD.loc[caliEnd-timedelta(days=caliYr*365.25):caliEnd]
    ret_cali_other_CAD = exRet_other_CAD.loc[caliEnd - timedelta(days=caliYr * 365.25):caliEnd]
    if ESG == 'Yes':
        ret_cali_ESG_equity_FI_USD = exRet_ESG_equity_FI_USD.loc[caliEnd - timedelta(days=caliYr * 365.25):caliEnd]

    # factors return for the calibration period
    ret_factor_cali = ret_factor.loc[caliEnd-timedelta(days=caliYr*365.25):caliEnd]

    currPrice = price_portf.loc[:caliEnd].iloc[-1]  # last observed price during calibration period
    testPrice = price_portf.loc[testStart:testEnd]  # rebalancing period prices
    testPeriod = len(price_portf.loc[testStart:testEnd])
    testDate2[t] = price_portf.loc[testStart:testEnd].index[-1].value  # last day during each rebalancing period

    if t == 0:
        # last observed portfolio value during current calibration period
        currentVal[t] = init_cap
    else:
        currentVal[t] = np.dot(currPrice, NoShares) + injection - trans_cost

        # store the current asset weights (before rebalance)
        x0[:, t] = (currPrice * NoShares) / currentVal[t]

    # portfolio construction
    # factor model: OLS calibration + optimization model, apply on each asset class
    # Risk-Parity
    if allocation_model == 'RP':
        if ESG == 'No':
            mu1, Q1, adj_R2_bar1 = OLS(ret_cali_equity_FI_USD, ret_factor_cali)
            x_equity_FI_USD[:, t] = RP(Q1) * countryPerc * equityFI_perc

            mu2, Q2, adj_R2_bar2 = OLS(ret_cali_other_USD, ret_factor_cali)
            x_other_USD[:, t] = RP(Q2) * countryPerc * (1-equityFI_perc)

            mu3, Q3, adj_R2_bar3 = OLS(ret_cali_equity_FI_CAD, ret_factor_cali)
            x_equity_FI_CAD[:, t] = RP(Q3) * countryPerc * equityFI_perc

            mu4, Q4, adj_R2_bar4 = OLS(ret_cali_other_CAD, ret_factor_cali)
            x_other_CAD[:, t] = RP(Q4) * countryPerc * (1-equityFI_perc)

        elif ESG == 'Yes':
            mu1, Q1, adj_R2_bar1 = OLS(ret_cali_equity_FI_USD, ret_factor_cali)
            x_equity_FI_USD[:, t] = RP(Q1) * (countryPerc * equityFI_perc - ESGperc)  # 20%

            mu2, Q2, adj_R2_bar2 = OLS(ret_cali_other_USD, ret_factor_cali)
            x_other_USD[:, t] = RP(Q2) * countryPerc * (1 - equityFI_perc)  # 10%

            mu3, Q3, adj_R2_bar3 = OLS(ret_cali_equity_FI_CAD, ret_factor_cali)
            x_equity_FI_CAD[:, t] = RP(Q3) * countryPerc * equityFI_perc  # 40%

            mu4, Q4, adj_R2_bar4 = OLS(ret_cali_other_CAD, ret_factor_cali)
            x_other_CAD[:, t] = RP(Q4) * countryPerc * (1 - equityFI_perc)  # 10%

            mu5, Q5, adj_R2_bar5 = OLS(ret_cali_ESG_equity_FI_USD, ret_factor_cali)
            x_ESG_equity_FI_USD[:, t] = RP(Q5) * ESGperc  # 20%
    # CVaR
    elif allocation_model == 'CVaR':
        if ESG == 'No':
            mu1, Q1, adj_R2_bar1 = OLS(ret_cali_equity_FI_USD, ret_factor_cali)
            x_equity_FI_USD[:, t] = CVaR(ret_cali_equity_FI_USD, alpha) * countryPerc * equityFI_perc

            mu2, Q2, adj_R2_bar2 = OLS(ret_cali_other_USD, ret_factor_cali)
            x_other_USD[:, t] = CVaR(ret_cali_other_USD, alpha) * countryPerc * (1 - equityFI_perc)

            mu3, Q3, adj_R2_bar3 = OLS(ret_cali_equity_FI_CAD, ret_factor_cali)
            x_equity_FI_CAD[:, t] = CVaR(ret_cali_equity_FI_CAD, alpha) * countryPerc * equityFI_perc

            mu4, Q4, adj_R2_bar4 = OLS(ret_cali_other_CAD, ret_factor_cali)
            x_other_CAD[:, t] = CVaR(ret_cali_other_CAD, alpha) * countryPerc * (1 - equityFI_perc)

        elif ESG == 'Yes':
            mu1, Q1, adj_R2_bar1 = OLS(ret_cali_equity_FI_USD, ret_factor_cali)
            x_equity_FI_USD[:, t] = CVaR(ret_cali_equity_FI_USD, alpha) * (countryPerc * equityFI_perc - ESGperc)  # 20%

            mu2, Q2, adj_R2_bar2 = OLS(ret_cali_other_USD, ret_factor_cali)
            x_other_USD[:, t] = CVaR(ret_cali_other_USD, alpha) * countryPerc * (1 - equityFI_perc)  # 10%

            mu3, Q3, adj_R2_bar3 = OLS(ret_cali_equity_FI_CAD, ret_factor_cali)
            x_equity_FI_CAD[:, t] = CVaR(ret_cali_equity_FI_CAD, alpha) * countryPerc * equityFI_perc  # 40%

            mu4, Q4, adj_R2_bar4 = OLS(ret_cali_other_CAD, ret_factor_cali)
            x_other_CAD[:, t] = CVaR(ret_cali_other_CAD, alpha) * countryPerc * (1 - equityFI_perc)  # 10%

            mu5, Q5, adj_R2_bar5 = OLS(ret_cali_ESG_equity_FI_USD, ret_factor_cali)
            x_ESG_equity_FI_USD[:, t] = CVaR(ret_cali_ESG_equity_FI_USD, alpha) * ESGperc  # 20%
    # Robust MVO
    elif allocation_model == 'robustMVO':
        if ESG == 'No':
            mu1, Q1, adj_R2_bar1 = OLS(ret_cali_equity_FI_USD, ret_factor_cali)
            x_equity_FI_USD[:, t] = robustMVO(mu1, Q1, riskAver, alpha) * countryPerc * equityFI_perc

            mu2, Q2, adj_R2_bar2 = OLS(ret_cali_other_USD, ret_factor_cali)
            x_other_USD[:, t] = robustMVO(mu2, Q2, riskAver, alpha) * countryPerc * (1 - equityFI_perc)

            mu3, Q3, adj_R2_bar3 = OLS(ret_cali_equity_FI_CAD, ret_factor_cali)
            x_equity_FI_CAD[:, t] = robustMVO(mu3, Q3, riskAver, alpha) * countryPerc * equityFI_perc

            mu4, Q4, adj_R2_bar4 = OLS(ret_cali_other_CAD, ret_factor_cali)
            x_other_CAD[:, t] = robustMVO(mu4, Q4, riskAver, alpha) * countryPerc * (1 - equityFI_perc)

        elif ESG == 'Yes':
            mu1, Q1, adj_R2_bar1 = OLS(ret_cali_equity_FI_USD, ret_factor_cali)
            x_equity_FI_USD[:, t] = robustMVO(mu1, Q1, riskAver, alpha) * (countryPerc * equityFI_perc - ESGperc)

            mu2, Q2, adj_R2_bar2 = OLS(ret_cali_other_USD, ret_factor_cali)
            x_other_USD[:, t] = robustMVO(mu2, Q2, riskAver, alpha) * countryPerc * (1 - equityFI_perc)

            mu3, Q3, adj_R2_bar3 = OLS(ret_cali_equity_FI_CAD, ret_factor_cali)
            x_equity_FI_CAD[:, t] = robustMVO(mu3, Q3, riskAver, alpha) * countryPerc * equityFI_perc

            mu4, Q4, adj_R2_bar4 = OLS(ret_cali_other_CAD, ret_factor_cali)
            x_other_CAD[:, t] = robustMVO(mu4, Q4, riskAver, alpha) * countryPerc * (1 - equityFI_perc)

            mu5, Q5, adj_R2_bar5 = OLS(ret_cali_ESG_equity_FI_USD, ret_factor_cali)
            x_ESG_equity_FI_USD[:, t] = robustMVO(mu5, Q5, riskAver, alpha) * ESGperc
    # adjusted R-square: check model fitting
    if ESG == 'No':
        adj_R2_bar[t] = (adj_R2_bar1 + adj_R2_bar2 + adj_R2_bar3 + adj_R2_bar4) / 4
        # new asset weight
        x[:, t] = np.append(np.append(np.append(x_equity_FI_USD[:, t], x_other_USD[:, t]), x_equity_FI_CAD[:, t]),
                            x_other_CAD[:, t])
        tickers = tickers_nonESG_equity_FI_USD + tickers_nonESG_other_USD + tickers_nonESG_equity_FI_CAD +                   tickers_nonESG_other_CAD
        # weight dataframe
        weights = pd.DataFrame(x.T, columns=tickers)
    elif ESG == 'Yes':
        adj_R2_bar[t] = (adj_R2_bar1 + adj_R2_bar2 + adj_R2_bar3 + adj_R2_bar4 + adj_R2_bar5) / 5
        x[:, t] = np.append(np.append(np.append(np.append(x_equity_FI_USD[:, t], x_other_USD[:, t]),
                                                x_equity_FI_CAD[:, t]), x_other_CAD[:, t]), x_ESG_equity_FI_USD[:, t])
        tickers = tickers_nonESG_equity_FI_USD + tickers_nonESG_other_USD + tickers_nonESG_equity_FI_CAD +                   tickers_nonESG_other_CAD + tickers_ESG_equity_USD
        weights = pd.DataFrame(x.T, columns=tickers)

    NoShares_new = x[:, t] * currentVal[t] / currPrice  # Number of shares to be allocated to each ETFs
    trans_cost = sum(abs(NoShares_new - NoShares) * cost)
    NoShares = NoShares_new

    # calculate the  portfolio value at each month during the test period
    for j in range(testPeriod):
        index_portf = np.where(exRet_portf.loc[startDate:].index == testStart)[0][0]
        # loan interest accumulation
        rf_7yr = 1.56  # used 7-year treasury rate from 2016-04-01 which is 1.56%
        loan_pay = loan * (1 + rf_7yr / 100) ** ((index_portf+j)/12)  # accumulated by 7-year treasury rate

        if j == 0:
            price_portf.loc[testStart]
            portfValue[index_portf+j] = np.dot(testPrice, NoShares_new)[j] - loan_pay - trans_cost
        else:
            portfValue[index_portf+j] = np.dot(testPrice, NoShares_new)[j] - loan_pay

    # calculate the turnover rate
    if t > 0:
        turnover[t] = sum(abs(x[:, t] - x0[:, t]))

    # Update the calibration and out-of-sample test periods
    if t < 9:
        testStart = exRet_portf.loc[testStart:].index[6]  # each rebalancing period start date
        testEnd = exRet_portf.loc[testStart:].index[6] - timedelta(days=1)  # each rebalancing period end date
        caliEnd = testStart - timedelta(days=1)  # the end of each calibration period
    elif t == 9:
        testStart = exRet_portf.loc[testStart:].index[6]  # each rebalancing period start date
        testEnd = exRet_portf.loc[testStart:].index[-1]  # each rebalancing period end date
        caliEnd = testStart - timedelta(days=1)  # the end of each calibration period


# outputs:
print('portfolio value evolution:' + str(portfValue.T))
print(weights)  # portfolio weighting
# print('portfolio weighting:' + str(x.T))
print('turnover:' + str(turnover))
print('adjusted R-squared:' + str(adj_R2_bar))
print('date:' + str(testDate))

# Check the correlation heatmap for our dataset
fig, ax = plt.subplots(figsize=(10,10))
corr = exRet_portf.corr()
sns.heatmap(corr, xticklabels=corr.columns, yticklabels=corr.columns, linewidths=.1, vmin=-1, vmax=1)
ax.set_title('Heatmap for all ETFs excess returns', fontsize=20)
plt.show()


# ---------------- system implementation end ----------------------

# ### implement performance and risk analysis below

# In[10]:


# remove injection from cashflow
portfValue_wocf = portfValue.copy()
portfPnL_wocf = np.diff(portfValue_wocf)
for i in range(len(portfPnL_wocf)):
    if i%injectFreq==injectFreq-1:
        portfPnL_wocf[i] = portfPnL_wocf[i] - injection
portfRet_wocf = portfPnL_wocf/portfValue_wocf[:-1]


# ---------------- construct benchmark ---------------

# In[20]:


# Link to the benchmark data
df_bm = pd.read_excel(dir + r'\Benchmark.xlsx', sheet_name='Prices', index_col=0,engine='openpyxl')
df_bm_val = df_bm.loc[testDate[0:]].copy()
df_bm = df_bm.pct_change()
df_bm = df_bm.loc[testDate[1:]]
# Prepare the risk free data
rf = rf.loc[testDate[1:]]


# In[21]:


# calculate the weight of each asset class for benchmark indices
# benchmark strategy: use the initial weight, buy and hold
x_equity = np.sum(x_equity_FI_USD[:5,0])+np.sum(x_equity_FI_CAD[0,0])+np.sum(x_ESG_equity_FI_USD[:,0])
x_fi = np.sum(x_equity_FI_USD[5:,0])+np.sum(x_equity_FI_CAD[1:,0])
x_REIT = np.sum(x_other_USD[0,0])+np.sum(x_other_CAD[:,0])
x_Commodity = np.sum(x_other_USD[1:,0])
ini_weight = [x_equity,x_fi,x_REIT,x_Commodity]
weighted_total_return = []
weighted_total_val = []
for i in range(df_bm.shape[0]):
    total = np.dot(df_bm.values.tolist()[i],ini_weight)
    weighted_total_return.append(total)
df_bm["Weighted Total Return"] = weighted_total_return

for i in range(df_bm_val.shape[0]):
    total_val = np.dot(df_bm_val.values.tolist()[i],ini_weight)
    weighted_total_val.append(total_val)
df_bm_val["Weighted Total Value"] = weighted_total_val


# In[22]:


df_bm["Weighted Total Return"]


# --------------------sensitivity analysis-------------------

# In[11]:


# handling df_macro
df_macro = pd.read_excel(dir + r'\Macro_Data.xlsx', sheet_name='Index', index_col=0, engine='openpyxl')
df_macro = df_macro.drop(['AAA', 'BAA'], axis=1)
# df_macro['USACPIALLMINMEI'] = df_macro['USACPIALLMINMEI'].pct_change() * 100
# df_macro['INDPRO'] = df_macro['INDPRO'].pct_change() * 100
# df_macro[['MCOILWTICO', 'HOUST']] = df_macro[['MCOILWTICO', 'HOUST']].pct_change() * 100
df_macro[['INDPRO', 'USACPIALLMINMEI', 'CBS', 'T10Y3MM', 'MCOILWTICO', 'HOUST', 'UNRATE']] = df_macro[['INDPRO', 'USACPIALLMINMEI', 'CBS', 'T10Y3MM', 'MCOILWTICO', 'HOUST', 'UNRATE']].pct_change() * 100

macro_corr = df_macro[['GS30', 'GS20', 'GS10', 'GS5', 'GS3', 'GS1', 'GS3M']].corr()
sns.heatmap(macro_corr, xticklabels=macro_corr.columns, yticklabels=macro_corr.columns, annot=True,linewidths=.1, vmin=-1, vmax=1)
ax.set_title('Heatmap for the yield curve', fontsize=20)
plt.show()

# PCA the yield curve
yieldcurve = df_macro[['GS30', 'GS20', 'GS10', 'GS5', 'GS3', 'GS1', 'GS3M']]
yieldcurve = StandardScaler().fit_transform(yieldcurve)
pca = PCA(n_components=1)
principalComponents = pca.fit_transform(yieldcurve)
principalDf = pd.DataFrame(data=principalComponents, columns=['principal component'])
df_macro['Yield Curve PCA'] = principalDf.values
df_macro = df_macro.drop(['GS30', 'GS20', 'GS10', 'GS5', 'GS3', 'GS1', 'GS3M'], axis=1)
# df_macro['Yield Curve PCA']=df_macro['Yield Curve PCA'].pct_change() * 100
df_macro=df_macro.iloc[1:]


# In[3]:


# We tried different combinations of the economic factors for the regression with market factors
# and checked the stats like R-adjusted square and P-values for each factor. Finally, we decided to choose the
# following economic factors for the stress testing model.

# We tried standard scaler to normalize the macro dataframe,
# however the performance is not improving.

# We use backward elimination approach, where all variables are initially included, 
# and in each step, the most statistically insignificant variable is dropped until the R adjusted square stop increasing


# In[13]:


df_macro_new = df_macro[['USACPIALLMINMEI', 'CBS', 'T10Y3MM', 'MCOILWTICO', 'HOUST',
       'UNRATE', 'Yield Curve PCA']]


# In[14]:


# sensitivity analysis for the allocation model
portfRet_up=np.zeros((len(df_macro_new.columns),len(df_macro_new.columns)+2))
portfRet_down=np.zeros((len(df_macro_new.columns),len(df_macro_new.columns)+2))
portfRet_wocf = pd.DataFrame(portfRet_wocf,index=testDate[1:],columns=['portfValue'])

# macro data for the test period
df_macro_test = df_macro_new.loc[portfRet_wocf.index]
# fit model for x=macro factor returns, y=portofolio return during the test period
df_macro_test = sm.add_constant(df_macro_test)
model = sm.OLS(portfRet_wocf, df_macro_test).fit()


# In[15]:


model.summary()


# In[16]:


# different scenarios for the macro factor return

# normal scenario
norm_scenario = df_macro_test.median()
portfRet_norm = np.dot(np.array(norm_scenario), np.array(model.params))
print('For the normal scenario, the monthly portfolio return is {}'.format(portfRet_norm))

# up scenario for each macro factor - shocking factor return one by one
for i in range(norm_scenario.shape[0]):
    if i != 0:
        norm_scenario_up = norm_scenario.copy()
        macro_fact = norm_scenario_up.index[i]
        norm_scenario_up[macro_fact]=norm_scenario_up[macro_fact]*2
        portfRet_up[i-1,:-1] = norm_scenario_up
        portfRet_up[i-1,-1] = np.dot(np.array(norm_scenario_up), np.array(model.params))
portfRet_up_df = pd.DataFrame(portfRet_up,columns = norm_scenario.index.to_list() + ['Monthly Return'])
portfRet_up_df['AllocationModel'] = allocation_model

# down scenario for each macro factor
for i in range(norm_scenario.shape[0]):
    if i != 0:
        norm_scenario_down = norm_scenario.copy()
        macro_fact = norm_scenario_down.index[i]
        norm_scenario_down[macro_fact]=norm_scenario_down[macro_fact]*0.5
        portfRet_down[i-1,:-1] = norm_scenario_down
        portfRet_down[i-1,-1] = np.dot(np.array(norm_scenario_down), np.array(model.params))
portfRet_down_df = pd.DataFrame(portfRet_down,columns = norm_scenario.index.to_list() + ['Monthly Return'])
portfRet_down_df['AllocationModel'] = allocation_model

# Method2: use copula to simulate macro factor returns
copula = VineCopula('regular')
copula.fit(df_macro_test.drop(['const'],axis=1))
samples = copula.sample(1000)
samples.insert(0, 'constant', 1)
# Use the simulated factor returns to generate portfolio return using fitted OLS parameters
samplePortfolio = np.dot(np.array(samples), np.array(model.params))
samples['Portfolio Estimated Return'] = samplePortfolio
# best 3 up scenario and the best 3 down scenario for the simulated factor returns
upScenario1 = samples.sort_values('Portfolio Estimated Return', ascending=0).iloc[:3, :]
downScenario1 = samples.sort_values('Portfolio Estimated Return', ascending=1).iloc[:3, :]
upScenario1['AllocationModel'] = allocation_model
downScenario1['AllocationModel'] = allocation_model

print('the best 3 up scenario: {}'.format(upScenario1))
print('the best 3 down scenario: {}'.format(downScenario1))


# In[ ]:


# formatting the 'sensitivity.xlsx' output file
if allocation_model == 'RP':
    arow=1
elif allocation_model == 'robustMVO':
    arow=10
elif allocation_model == 'CVaR':
    arow=20

# Export results to excel
from openpyxl import load_workbook
pd.read_excel(dir + r'\ESG_Prices.xlsx', sheet_name='Prices', index_col=0, engine='openpyxl')
book = load_workbook(dir + r'\sensitivity.xlsx')
writer = pd.ExcelWriter(dir + r'\sensitivity.xlsx', engine='openpyxl') 
writer.book = book

## ExcelWriter for some reason uses writer.sheets to access the sheet.
## If you leave it empty it will not know that sheet Main is already there
## and will create a new sheet.

writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

portfRet_up_df.to_excel(writer,
                        sheet_name='portfRet_up_df',startrow=arow)
portfRet_down_df.to_excel(writer,
                          sheet_name='portfRet_down_df',startrow=arow)
upScenario1.to_excel(writer,
                     sheet_name='upScenario',startrow=arow)
downScenario1.to_excel(writer,
                       sheet_name='downScenario',startrow=arow)

writer.save()
writer.close()


# --------------------- sensitivity test for benchmark ----------------------

# In[24]:


# this is all macro factors that we select
df_macro_bm = df_macro[['USACPIALLMINMEI', 'T10Y3MM', 'MCOILWTICO', 
       'UNRATE']]


# In[25]:


# sensitivity analysis for the allocation model
bmRet_up=np.zeros((len(df_macro_bm.columns),len(df_macro_bm.columns)+2))
bmRet_down=np.zeros((len(df_macro_bm.columns),len(df_macro_bm.columns)+2))

# macro data for the test period
df_macro_bm = df_macro_bm.loc[df_bm["Weighted Total Return"].index]
# fit model for x=macro data, y=portofolio value during the test period
df_macro_bm = sm.add_constant(df_macro_bm)
model = sm.OLS(df_bm["Weighted Total Return"], df_macro_bm).fit()


# In[26]:


model.summary()


# ------------------------ scenario analysis ------------------------

# In[19]:


# Scenario Analysis - Stress VaR
# factors
factors_all = pd.read_csv(dir + r'\FF_5_Factor_all.csv', index_col=0,parse_dates=True)
ret_factor_all = factors_all[['Mkt-RF','SMB','HML','RMW','CMA']].iloc[:-1,:]
ret_factor_test = ret_factor_all.loc[exRet_portf.index] # exRet_portf.index is from 2007-02-01 to 2016-03-01

# historical asset returns
data_all = pd.read_excel(dir + r'\Historical_Data.xlsx', sheet_name='Sheet1', index_col=0,engine='openpyxl')
data_all = data_all.pct_change()

# Calibrate OLS model on the current dataset: X=famma french 5 factor returns, y=asset excess return
scenario_coef = pd.DataFrame(columns=weights.columns)
exRet_portf_test = exRet_portf.copy()
ret_factor_test = sm.add_constant(ret_factor_test)

for tickerindex in range(len(exRet_portf_test.columns)):
    ticker = exRet_portf_test.columns[tickerindex]
    scenario_model = sm.OLS(exRet_portf_test[ticker], ret_factor_test).fit()
    scenario_coef[ticker] = scenario_model.params

# use the results from calibrated OLS model to simulate historical asset returns
ret_factor_all = sm.add_constant(ret_factor_all)
regressed = np.dot(ret_factor_all,scenario_coef)
regressed = pd.DataFrame(regressed,columns = exRet_portf.columns, index=ret_factor_all.index)

# trim data_all with only useful ticker left
data_all = data_all[exRet_portf_test.columns]

# if historical dataset is empty, replace it with the simulated past asset return
for i in range(data_all.shape[0]): #iterate over rows
    for j in range(data_all.shape[1]): #iterate over columns
        adate = data_all.index[i]
        aticker = data_all.columns[j]
        if pd.isna(data_all.iloc[i, j]):
            data_all.loc[adate,aticker] = regressed.loc[adate,aticker]

# list of historical scenarios
# using the weights during the initial test period

# all historical data are simulated so it's not very reliable. skip it
# factors_AsianCrisis = ret_factor_all.loc['1998-04-01':'1998-10-01']
# AsianCrisis = np.dot(np.dot(factors_AsianCrisis,scenario_coef),weights.iloc[0])
# print('For {}:'.format('AsianCrisis'))
# VaRCalc(AsianCrisis)

# this extra weights[data_all.columns] is to align the column of data_all and weights.
techBubble = np.dot(data_all.loc['2000-03-01':'2002-09-01'],weights[data_all.columns].iloc[-1])
print('For {}:'.format('techBubble'))
VaRCalc(techBubble)

sellOff911 = np.dot(data_all.loc['2001-07-01':'2001-09-01'],weights[data_all.columns].iloc[-1])
print('For {}:'.format('sellOff911'))
VaRCalc(sellOff911)

subprimeRecession = np.dot(data_all.loc['2008-01-01':'2009-06-01'],weights[data_all.columns].iloc[-1])
print('For {}:'.format('subprimeRecession'))
VaRCalc(subprimeRecession)

Summer2011 = np.dot(data_all.loc['2011-06-01':'2011-10-01'],weights[data_all.columns].iloc[-1])
print('For {}:'.format('Summer2011'))
VaRCalc(Summer2011)


# -------------------Portfolio Performance and risk Measure -------------------

# In[28]:


# Calculate Kurtosis and Skewness
kurtosis = stats.kurtosis(portfRet_wocf)
skewness = stats.skew(portfRet_wocf)
print("Kurtosis:", round(kurtosis[0],5))
print("Skewness:", round(skewness[0],5))

# Export results to excel
import openpyxl
wb = openpyxl.load_workbook(dir + r'\Report.xlsx')
sheet = wb['Report']
sheet['C22'].value = round(kurtosis[0],5)
sheet['C23'].value = round(skewness[0],5)
sheet['C5'].value = startDate.date()
sheet['C6'].value = reportDate.date()

wb.save(dir + r'\Report.xlsx')


# In[29]:


# Calculate the Return, Volatility, Correlation with Benchmark, Information Ratio and Sharpe Ratio

TotalReturn = np.prod(1+portfRet_wocf.iloc[:,0])-1 # Total Portfolio Return
annualret=(np.prod(1+portfRet_wocf.iloc[:,0])**(1/len(portfRet_wocf.iloc[:,0]))-1)*12 # Annualized Portfolio Return
annualret_1=(np.prod(1+portfRet_wocf.iloc[-12:,0])**(1/len(portfRet_wocf.iloc[-12:,0]))-1)*12 # 1-yr Annualized Return
annualret_3=(np.prod(1+portfRet_wocf.iloc[-36:,0])**(1/len(portfRet_wocf.iloc[-36:,0]))-1)*12 # 3-yr Annualized Return
annualret_5=(np.prod(1+portfRet_wocf.iloc[-60:,0])**(1/len(portfRet_wocf.iloc[-60:,0]))-1)*12 # 5-yr Annualized Return
stdret= np.std(portfRet_wocf.iloc[:,0], ddof=1)*np.sqrt(12) # Annualized Portfolio Standard Deviation
Corr = np.corrcoef(portfRet_wocf.iloc[:,0], df_bm.iloc[:,-1])[0][1] # Correlation between Portfolio and Weighted Benchmark
Sharpe = (annualret-0.0156)/stdret # Annualized Sharpe Ratio
TotalBenchmarkReturn = np.prod(1+df_bm.iloc[:,-1])-1 # Total Weighted Benchmark Return
annualbret = (np.prod(1+df_bm.iloc[:,-1])**(1/len(df_bm.iloc[:,-1]))-1)*12 # Annualized Benchmark Return
TrackingError = np.std(portfRet_wocf.iloc[:,0]-df_bm.iloc[:,-1])*np.sqrt(12) # Tracking Error
InfoRatio = (annualret-annualbret)/TrackingError # Annualized Information Ratio

print("Annualized Return", round(annualret,2))
print("1-yr Annualized Return", round(annualret_1,2))
print("3-yr Annualized Return", round(annualret_3,2))
print("5-yr Annualized Return", round(annualret_5,2))
print("Annualized Volatility", round(stdret,2))
print("Correlation with Benchmark", round(Corr,2))
print("Tracking Error", round(TrackingError,2))
print("Information Ratio", round(InfoRatio,2))
print("Sharpe Ratio", round(Sharpe,2))

# Export results to excel
wb = openpyxl.load_workbook(dir + r'\Report.xlsx')
sheet = wb['Report']
sheet['C8'].value = round(annualret,4)
sheet['C9'].value = round(annualret_1,4)
sheet['C10'].value = round(annualret_3,4)
sheet['C11'].value = round(annualret_5,4)
sheet['C12'].value = round(stdret,4)
sheet['C17'].value = round(Corr,2)
sheet['C18'].value = round(InfoRatio,4)
sheet['C13'].value = round(Sharpe,4)
sheet['C19'].value = round(TrackingError,2)

wb.save(dir + r'\Report.xlsx')


# In[30]:


#Calculate Max Drawdown
import quantstats as qs
max_drawdown=qs.stats.max_drawdown(portfRet_wocf.iloc[:,0])
print("max_drawdown", round(max_drawdown,2))

# Export results to excel
wb = openpyxl.load_workbook(dir + r'\Report.xlsx')
sheet = wb['Report']
sheet['C20'].value = round(max_drawdown,4)

wb.save(dir + r'\Report.xlsx')


# In[31]:


# Calculate VaR and CVaR
VaR_1, value_1, CVaR_1, value_C_1, VaR_2, value_2, CVaR_2, value_C_2 = VaRCalc(portfRet_wocf.iloc[:,0])

# Export results to excel
wb = openpyxl.load_workbook(dir + r'\Report.xlsx')
sheet = wb['Report']
str(round(annualret,4) * 100) + '%'
sheet['C25'].value = round(VaR_1,4)
sheet['C26'].value = round(CVaR_1,4)
sheet['C27'].value = round(VaR_2,4)
sheet['C28'].value = round(CVaR_2,4)
sheet['D25'].value = round(value_1,2)
sheet['D26'].value = round(value_C_1,2)
sheet['D27'].value = round(value_2,2)
sheet['D28'].value = round(value_C_2,2)

wb.save(dir + r'\Report.xlsx')


# In[32]:


#Measure Beta and alpha
(beta, alpha) = stats.linregress(list(df_bm.iloc[:, -1]), list(portfRet_wocf.iloc[:,0].dropna()))[0:2]
print ("Beta:", round(beta, 2))
print ("Alpha:", round(alpha*12*100, 2))

# Export results to excel
wb = openpyxl.load_workbook(dir + r'\Report.xlsx')
sheet = wb['Report']
sheet['C14'].value = round(alpha*12*100, 2)
sheet['C15'].value = round(beta, 2)

wb.save(dir + r'\Report.xlsx')


# In[33]:


# Plot return of portfolio vs return of benchmark
plt.figure(figsize=(8,5.3))
plt.plot(portfRet_wocf.iloc[:,0].cumsum()*100,label="Portfolio")
plt.plot(df_bm.iloc[:,-1].cumsum()*100,label="Benchmark")
plt.xlabel("Years")
plt.ylabel("Returns")
plt.legend()
plt.title("Performance of Portfolio vs Benchmark")
plt.savefig('return.png')
plt.show()

# Export to Excel
wb = openpyxl.load_workbook(dir + r'\Report.xlsx')
ws = wb['Report']
img = openpyxl.drawing.image.Image(dir + r'\return.png')
img.anchor = 'F31' # Or whatever cell location you want to use.
ws.add_image(img)
wb.save(dir + r'\Report.xlsx')


# In[34]:


# Plot PnL of portfolio vs PnL of benchmark
pnl_portf = np.append((portfValue_wocf[0] - cap), portfRet_wocf * portfValue_wocf[0]).cumsum()
pnl_bm = np.append((portfValue_wocf[0] - cap), df_bm.iloc[:,-1]*portfValue_wocf[0]).cumsum()
plt.figure(figsize=(7.5,5.3))
plt.plot(testDate, pnl_portf,label="Portfolio")
plt.plot(testDate, pnl_bm,label="Benchmark")
plt.xlabel("Years")
plt.ylabel("Returns")
plt.legend()
plt.title("PnL of Portfolio vs Benchmark")
plt.savefig('PnL.png')
plt.show()

# Export to Excel
wb = openpyxl.load_workbook(dir + r'\Report.xlsx')
ws = wb['Report']
img = openpyxl.drawing.image.Image(dir + r'\PnL.png')
img.anchor = 'B31' # Or whatever cell location you want to use.
ws.add_image(img)
wb.save(dir + r'\Report.xlsx')


# In[35]:


# Portfolio Value Evolution
plt.figure(figsize=(7.5,5.3))
plt.plot(testDate, portfValue,label="Portfolio")
plt.xlabel("Years")
plt.ylabel("Returns")
plt.legend()
plt.title("Portfolio Value Evolution")
plt.savefig('portfvalue.png')
plt.show()

# Export to Excel
wb = openpyxl.load_workbook(dir + r'\Report.xlsx')
ws = wb['Report']
img = openpyxl.drawing.image.Image(dir + r'\portfvalue.png')
img.anchor = 'Q8' # Or whatever cell location you want to use.
ws.add_image(img)
wb.save(dir + r'\Report.xlsx')


# In[36]:


# Use quantstats to plot performance analysis - for reference
qs.plots.yearly_returns(portfRet_wocf.iloc[:, 0],df_bm.iloc[:, -1], figsize=(7.5,5.2),
                        savefig={'fname': dir + r'\EOY_return.png'}) # EOY Returns
qs.plots.drawdown(portfRet_wocf.iloc[:, 0], figsize=(12, 6), savefig={'fname': dir + r'\drawdown.png'})  # drawdowns
qs.plots.histogram(portfRet_wocf.iloc[:, 0], figsize=(7.5,5.2), savefig={'fname': dir + r'\histogram.png'})  # return distribution
qs.plots.monthly_heatmap(portfRet_wocf.iloc[:, 0], figsize=(10, 5), 
                         savefig={'fname': dir + r'\monthly_heatmap.png'})  # monthly return heatmap

# Export to Excel
wb = openpyxl.load_workbook(dir + r'\Report.xlsx')
ws = wb['Report']

img = openpyxl.drawing.image.Image(dir + r'\EOY_return.png')
img.anchor = 'Q27' # Or whatever cell location you want to use.
ws.add_image(img)

img = openpyxl.drawing.image.Image(dir + r'\drawdown.png')
img.anchor = 'S44' # Or whatever cell location you want to use.
ws.add_image(img)

img = openpyxl.drawing.image.Image(dir + r'\histogram.png')
img.anchor = 'Z27' # Or whatever cell location you want to use.
ws.add_image(img)

img = openpyxl.drawing.image.Image(dir + r'\monthly_heatmap.png')
img.anchor = 'C46' # Or whatever cell location you want to use.
ws.add_image(img)

wb.save(dir + r'\Report.xlsx')


# In[39]:


# Plot the weight evolution

# Separate into different classes based on currencies, ESG or Non-ESG, and assets and get the corresponding weights
tickers_nonESG_equity_USD = ['XNTK', 'KCE', 'XLE', 'XLV', 'XLU']
tickers_nonESG_FI_USD = ['SHY', 'TLT', 'TIP', 'LQD', 'MHWIX']
tickers_nonESG_REIT_USD = ['IYR']
tickers_nonESG_Commodity_USD = [ 'GLD', 'GSG']
tickers_nonESG_equity_CAD = ['XIT.TO']
tickers_nonESG_FI_CAD = ['XBB.TO', 'XRB.TO']
tickers_nonESG_REIT_CAD = ['XRE.TO']
tickers_ESG_equity_USD = ['XLB', 'SOXX']

weights_asset_class = pd.DataFrame() # create a new dataframe to store the weight for each asset class
weights_asset_class["nonESG_Equity_USD"] = weights[tickers_nonESG_equity_USD].sum(axis=1)
weights_asset_class["nonESG_FI_USD"] = weights[tickers_nonESG_FI_USD].sum(axis=1)
weights_asset_class["nonESG_REIT_USD"] = weights[tickers_nonESG_REIT_USD].sum(axis=1)
weights_asset_class["nonESG_Commodity_USD"] = weights[tickers_nonESG_Commodity_USD].sum(axis=1)
weights_asset_class["nonESG_Equity_CAD"] = weights[tickers_nonESG_equity_CAD].sum(axis=1)
weights_asset_class["nonESG_FI_CAD"] = weights[tickers_nonESG_FI_CAD].sum(axis=1)
weights_asset_class["nonESG_REIT_CAD"] = weights[tickers_nonESG_REIT_CAD].sum(axis=1)
weights_asset_class["ESG_Equity_USD"] = weights[tickers_ESG_equity_USD].sum(axis=1)
weights_asset_class=weights_asset_class.set_index(testDate2) # set the index to be rebalancing dates

x = weights_asset_class.plot.line(figsize=(8,5.5), title = "Weight Evolution by Asset Class & Currency", xlabel = "Rebalancing Periods")
x.legend(loc='center left', bbox_to_anchor=(0.5, 0.75), prop={'size': 7})
x.figure.savefig('weight.png')
print(x)

# Export to Excel
wb = openpyxl.load_workbook(dir + r'\Report.xlsx')
ws = wb['Report']
img = openpyxl.drawing.image.Image(dir + r'\weight.png')
img.anchor = 'Z8' # Or whatever cell location you want to use.
ws.add_image(img)
wb.save(dir + r'\Report.xlsx')


# In[40]:


# portfolio weighting 
weights_asset_class.index = [''] * 11
weights_asset_class.iloc[-1,:].plot.pie(autopct='%.2f',fontsize=8, figsize=(14, 7),subplots=True)
plt.title("Current Portfolio Weighting",fontsize=13)
plt.legend(loc='upper left', bbox_to_anchor=(1, 0.5), prop={'size': 7})
plt.savefig('weighting_pie.png')
plt.show()

# Export to Excel
wb = openpyxl.load_workbook(dir + r'\Report.xlsx')
ws = wb['Report']
img = openpyxl.drawing.image.Image(dir + r'\weighting_pie.png')
img.anchor = 'C5' # Or whatever cell location you want to use.
ws.add_image(img)
wb.save(dir + r'\Report.xlsx')


# In[42]:


#Risk Attribution
price_portf_1y = price_portf.loc[testDate]  # ETF price for the test period
price_portf_1y = price_portf_1y.pct_change()  # ETF monthly return for the test period
price_portf_1y = price_portf_1y.iloc[1:]
cov_matrix_1y = price_portf_1y.cov()
cov_matrix_1y_mtx = cov_matrix_1y.to_numpy()  # ETF monthly return variance-covariance matrix
latest_weight = weights.iloc[-1]  # the weight for each asset of the last rebalancing period
# calculate the risk contribution of each ETF to the portfolio risk
risk_contribution = calculate_risk_contribution(latest_weight, cov_matrix_1y_mtx)  
risk_contribution = risk_contribution / np.sum(calculate_risk_contribution(latest_weight, cov_matrix_1y_mtx))
risk_contribution = pd.DataFrame(np.array(risk_contribution).T, columns = tickers)
# seperate the risk contribution into asset classes
rc_asset_class = pd.DataFrame()
rc_asset_class["nonESG_Equity_USD"] = risk_contribution[tickers_nonESG_equity_USD].sum(axis=1)
rc_asset_class["nonESG_FI_USD"] = risk_contribution[tickers_nonESG_FI_USD].sum(axis=1)
rc_asset_class["nonESG_REIT_USD"] = risk_contribution[tickers_nonESG_REIT_USD].sum(axis=1)
rc_asset_class["nonESG_Commodity_USD"] = risk_contribution[tickers_nonESG_Commodity_USD].sum(axis=1)
rc_asset_class["nonESG_Equity_CAD"] = risk_contribution[tickers_nonESG_equity_CAD].sum(axis=1)
rc_asset_class["nonESG_FI_CAD"] = risk_contribution[tickers_nonESG_FI_CAD].sum(axis=1)
rc_asset_class["nonESG_REIT_CAD"] = risk_contribution[tickers_nonESG_REIT_CAD].sum(axis=1)
rc_asset_class["ESG_Equity_USD"] = risk_contribution[tickers_ESG_equity_USD].sum(axis=1)

print ("###################Risk Attribution###################")
riskAttribution = rc_asset_class.T
print (round(riskAttribution*100,2))
riskAttribution.columns=['']
riskAttribution[''].plot.pie(autopct='%.2f', fontsize=12, figsize=(16, 8),subplots=True)
plt.title("Risk Attribution",fontsize=25)
plt.legend(loc='upper left', bbox_to_anchor=(1, 0.5), prop={'size': 9})
plt.savefig('risk_attribution.png')
plt.show()

# Export to Excel
wb = openpyxl.load_workbook(dir + r'\Report.xlsx')
ws = wb['Report']
img = openpyxl.drawing.image.Image(dir + r'\risk_attribution.png')
img.anchor = 'AL24' # Or whatever cell location you want to use.
ws.add_image(img)
wb.save(dir + r'\Report.xlsx')

book = load_workbook(dir + r'\Report.xlsx')
writer = pd.ExcelWriter(dir + r'\Report.xlsx', engine='openpyxl') 
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
round(riskAttribution*100,2).to_excel(writer, sheet_name='Report',startcol=36,startrow=29)
writer.save()


# In[43]:


#Return Attribution Version 2 - consider injection
df_pfv = pd.DataFrame(portfValue_wocf, index = testDate) # change array of portfolio values to be dataframe with index of test dates
df_pfv_rb = df_pfv.loc[testDate2] # retreive the portfolio values of rebalancing dates
weights = weights.set_index(testDate2) # set the index of weights to be rebalancing dates
df_pfv_asset = weights.mul(df_pfv_rb.iloc[:,0], axis=0) # get the value of each asset class based on the weights
tickerAll = tickers_nonESG_equity_USD + tickers_nonESG_FI_USD + tickers_nonESG_REIT_USD + tickers_nonESG_Commodity_USD +  tickers_nonESG_equity_CAD + tickers_nonESG_FI_CAD + tickers_nonESG_REIT_CAD + tickers_ESG_equity_USD 
# for each rebalancing period, calculate the value change in all asset classes
valueAdd=[(df_pfv_asset.loc[testDate2[i]:testDate2[i+1],tickerAll].iloc[-1]-df_pfv_asset.loc[testDate2[i]:testDate2[i+1],tickerAll].iloc[0]) for i in range(len(testDate2)-1)]
valueAdd=pd.DataFrame(valueAdd)
# Calculate the sum of value changes of each asset class during all the rebalancing periods with injections removed
returnAttr=pd.Series()
returnAttr['nonESG_Equity_USD']=valueAdd[tickers_nonESG_equity_USD].sum().sum()
returnAttr['nonESG_FI_USD']=valueAdd[tickers_nonESG_FI_USD].sum().sum()
returnAttr['nonESG_REIT_USD']=valueAdd[tickers_nonESG_REIT_USD].sum().sum()
returnAttr['nonESG_Commodity_USD']=valueAdd[tickers_nonESG_Commodity_USD].sum().sum()
returnAttr['nonESG_Equity_CAD']=valueAdd[tickers_nonESG_equity_CAD].sum().sum()
returnAttr['nonESG_FI_CAD']=valueAdd[tickers_nonESG_FI_CAD].sum().sum()
returnAttr['nonESG_REIT_CAD']=valueAdd[tickers_nonESG_REIT_CAD].sum().sum()
returnAttr['ESG_Equity_USD']=valueAdd[tickers_ESG_equity_USD].sum().sum()
returnAttr["Return Attribution"]=returnAttr*(portfValue_wocf[-1]-100000-10000*10)/returnAttr.sum()

print ("###################Return Attribution###################")
df = returnAttr["Return Attribution"]
df= (round((df/df.sum())*100,2))
df=pd.DataFrame(df)
print (df)
df.columns=[""]
df[""].plot.pie(autopct='%.2f', fontsize=10, figsize=(16, 8))
plt.title("Return Attribution",fontsize=15)
plt.legend(loc='upper left', bbox_to_anchor=(1, 0.5), prop={'size': 7})
plt.savefig('return_attribution2.png')
plt.show()

# Export to Excel
wb = openpyxl.load_workbook(dir + r'\Report.xlsx')
ws = wb['Report']
img = openpyxl.drawing.image.Image(dir + r'\return_attribution2.png')
img.anchor = 'AL4' # Or whatever cell location you want to use.
ws.add_image(img)
wb.save(dir + r'\Report.xlsx')

book = load_workbook(dir + r'\Report.xlsx')
writer = pd.ExcelWriter(dir + r'\Report.xlsx', engine='openpyxl') 
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
df.to_excel(writer, sheet_name='Report',startc    ol=36,startrow=9)
writer.save()


# -------------------- automatically generate risk & return analysis into Excel dashboard --------------------

# In[44]:


from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

bottom_border = Border(bottom=Side(style='thin'))
l_border = Border(top=Side(style='thin'), left=Side(style='thin'))
r_border = Border(top=Side(style='thin'), right=Side(style='thin'))

font = Font(name='Calibri', size=11, bold=False, vertAlign=None, color='FF000000')
number_format = '0.00%'
date_format = 'yyyy-mm-dd'
fill = PatternFill("solid", fgColor="DDDDDD")

wb = openpyxl.load_workbook(dir + r'\Report.xlsx')
ws = wb['Report']
# property cell.border should be used instead of cell.style.border
#ws['B61'].border = thin_border
#ws['B62'].border = thin_border
#ws['B63'].border = thin_border
#ws['B63'].fill = fill
ws['C5'].number_format = date_format
ws['C6'].number_format = date_format
ws['C8'].number_format = number_format
ws['C9'].number_format = number_format
ws['C10'].number_format = number_format
ws['C11'].number_format = number_format
ws['C12'].number_format = number_format
ws['C13'].number_format = number_format
ws['C18'].number_format = number_format
ws['C19'].number_format = number_format
ws['C20'].number_format = number_format
ws['C25'].number_format = number_format
ws['C26'].number_format = number_format
ws['C27'].number_format = number_format
ws['C28'].number_format = number_format

ws['AK10'].border = l_border
ws['AL10'].border = r_border
ws.merge_cells('AK10:AL10')
ws['AK10'].alignment = Alignment(horizontal="center", vertical="center")

ws['AK30'].border = l_border
ws['AL30'].border = r_border
ws.merge_cells('AK30:AL30')
ws['AK30'].alignment = Alignment(horizontal="center", vertical="center")


wb.save(dir + r'\Report.xlsx')


# In[37]:


# rolling ratio

portfRet_wocf['rs'] = portfRet_wocf['portfValue'].rolling('182d').apply(my_rolling_sharpe)
portfRet_wocf['rv'] = portfRet_wocf['portfValue'].rolling('182d').apply(my_rolling_vol)

# rolling sharpe ratio (6 months)
portfRet_wocf['rs'].iloc[1:].plot(style='-', lw=3, color='indianred', label='Sharpe', figsize=[8,6])        .axhline(y = 0, color = "black", lw = 3)
plt.ylabel('Sharpe ratio')
plt.legend(loc='best')
plt.title('Rolling Sharpe ratio (6-month)')
fig.tight_layout()
plt.savefig('rolling_SR.png')
plt.show()

# Export to Excel
wb = openpyxl.load_workbook(dir + r'\Report.xlsx')
ws = wb['Report']
img = openpyxl.drawing.image.Image(dir + r'\rolling_SR.png')
img.anchor = 'AJ46' # Or whatever cell location you want to use.
ws.add_image(img)
wb.save(dir + r'\Report.xlsx')

# rolling volatility (6 months)
portfRet_wocf['rv'].iloc[1:].plot(style='-', lw=3, color='indianred', label='Sharpe', figsize=[8,6])        .axhline(y = 0, color = "black", lw = 3)
plt.ylabel('Volatility')
plt.legend(loc='best')
plt.title('Rolling Volatility (6-month)')
fig.tight_layout()
plt.savefig('rolling_vol.png')
plt.show()

# Export to Excel
wb = openpyxl.load_workbook(dir + r'\Report.xlsx')
ws = wb['Report']
img = openpyxl.drawing.image.Image(dir + r'\rolling_vol.png')
img.anchor = 'AR46' # Or whatever cell location you want to use.
ws.add_image(img)
wb.save(dir + r'\Report.xlsx')

